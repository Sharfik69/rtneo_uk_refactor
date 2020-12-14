import json
import os

import psycopg2
from openpyxl import Workbook

from colors import bcolors


class WeHave():
    def __init__(self, name, distr, cadastr):
        self.name = name
        self.distr = distr
        self.cadastr = cadastr
        self.conn = psycopg2.connect(dbname='reimport2', user='cuba',
                                     password='cuba', host='localhost')
        self.cursor = self.conn.cursor()

    def create_rtneo_file(self):
        print('\rФормируем и выполняем зарпос', end='')
        query = "select * from reimport_rosreestr_search left join reimport_rosreestr_object_info_full on reimport_rosreestr_search.cadastral_number = reimport_rosreestr_object_info_full.cadastral_number where reimport_rosreestr_object_info_full.cadastral_number like '{0}%' and address_notes like '%{1}%' and assignation_code not in ('204001000000') order by street, house, apartment"
        query = query.format(self.cadastr, self.distr)
        self.cursor.execute(query)
        print('\rЗапрос выполнен', end='')
        records = self.cursor.fetchall()
        wb = Workbook()
        s = wb.active
        row_ = 2
        header = ["Тип ук", "LicenseNumber", "LicenseRegDate", "FiasHouseGuid", "ContractGuid", "Описание района",
                  "Тип улицы", "Наименование улицы", "Дом", "Квартира", "Полный адрес", "Площадь", "Кадастровый номер",
                  "Код назначения", "Описание", "Количество владельцев", "Владельцы", "Количество дочерних", "Дочерние"]
        array = [22, 18, 23, 19, 35, 14, 34, 36]
        print('\rВыполняем запись в файл', end='')
        for id, head in enumerate(header):
            s.cell(row=1, column=id + 1).value = head

        for i in range(len(records)):
            full_info_obj = json.loads(records[i][33])

            district_info = '{0} {1}, {2} {3}'.format(
                full_info_obj.get("objectData", {}).get("address", {}).get("districtType", ''),
                full_info_obj.get("objectData", {}).get("address", {}).get("districtName", ''),
                full_info_obj.get("objectData", {}).get("address", {}).get("localityType", ''),
                full_info_obj.get("objectData", {}).get("address", {}).get("localityName", ''))

            s.cell(row=row_, column=6).value = district_info if district_info != ' ,  ' else ''

            for id, item in enumerate(array):
                if item == 22 and records[i][item] != None:
                    x = records[i][item].split('|')
                    if len(x) == 1:
                        x.append('УЛ')
                    s.cell(row=row_, column=7).value = x[1]
                    s.cell(row=row_, column=8 + id).value = x[0]
                elif item == 18 and records[i][item].endswith('||'):
                    s.cell(row=row_, column=8 + id).value = records[i][item].replace('||', '')
                else:
                    s.cell(row=row_, column=8 + id).value = records[i][item]

            row_ += 1
        print('\rСохраняем', end='')
        try:
            os.makedirs('Выгрузка/tmp/')
        except OSError:
            pass

        wb.save('Выгрузка/tmp/tmp.xlsx'.format(self.name))
        print(bcolors.OKGREEN + '\r1) ртнео файл выгружен' + bcolors.ENDC)
