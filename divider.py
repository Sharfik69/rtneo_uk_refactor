from openpyxl import Workbook, load_workbook

from colors import bcolors

def divider(column_specific, wb, daughter=True):
    s = wb['Sheet']
    assignation_code = {}
    i = 2
    while True:
        ass = s.cell(row=i, column=column_specific).value
        checker = s.cell(row=i, column=13).value == None and s.cell(row=i, column=19).value == None
        if checker:
            break
        if ass is None:
            i += 1
            continue
        if ass not in assignation_code:
            assignation_code[ass] = {}
            assignation_code[ass]['wb'] = Workbook()
            assignation_code[ass]['s'] = assignation_code[ass]['wb'].active
            assignation_code[ass]['row'] = 1
        for col_ in range(1, 30):
            assignation_code[ass]['s'].cell(row=assignation_code[ass]['row'], column=col_).value = s.cell(row=i,
                                                                                                          column=col_).value
        assignation_code[ass]['row'] += 1
        i += 1
        if not daughter:
            continue
        while s.cell(row=i, column=14).value is None:
            for col_ in range(19, 50):
                assignation_code[ass]['s'].cell(row=assignation_code[ass]['row'], column=col_).value = s.cell(row=i,
                                                                                                              column=col_).value
            assignation_code[ass]['row'] += 1
            i += 1
            if s.cell(row=i, column=19).value == None:
                break
    return assignation_code

def uk_divider(wb, daughter=True, name=''):
    s = wb['Sheet']
    i, our_row = 2, 1
    wb_new = Workbook()
    s_new = wb_new.active

    while True:
        checker = s.cell(row=i, column=13).value == None and s.cell(row=i, column=19).value == None
        if checker:
            break

        loc = s.cell(row=i, column=1).value
        if loc is None:
            i += 1
            continue

        for j in range(1, 40):
            s_new.cell(row=our_row, column=j).value = s.cell(row=i, column=j).value

        our_row += 1
        i += 1

        if not daughter:
            continue

        while s.cell(row=i, column=14).value is None:
            for col_ in range(19, 50):
                s_new.cell(row=our_row, column=col_).value = s.cell(row=i, column=col_).value
            our_row += 1
            i += 1

    wb_new.save('Выгрузка/' + name + '.xlsx')


class Divider():
    def __init__(self, wb=None, path='Выгрузка/2.xlsx'):
        if wb is not None:
            self.wb = wb
        else:
            self.wb = load_workbook(path)

    def divide_by_assignation_code(self):
        print('\rДелим на файлы по коду помещения', end='')
        assignation_code = divider(14, self.wb)
        for ass, item in assignation_code.items():
            item['wb'].save('Выгрузка/' + ass + '.xlsx')
        print(bcolors.OKGREEN + '\rФайл был разделен по коду жилого помещения' + bcolors.ENDC)

    def divide_by_type_uk(self):
        print('\rДелим на файлы по типу ук', end='')

        assignation_code = divider(1, self.wb)
        for ass, item in assignation_code.items():
            item['wb'].save('Выгрузка/' + ass + ' с дочерними.xlsx')

        assignation_code = divider(1, self.wb, daughter=False)
        for ass, item in assignation_code.items():
            item['wb'].save('Выгрузка/' + ass + ' без дочерних.xlsx')

        print(bcolors.OKGREEN + '\rФайл был разделен по типу ук' + bcolors.ENDC)

    def only_uk(self):
        print('\rДелим на файлы только с ук', end='')

        uk_divider(self.wb, name='ук с дочерними')
        uk_divider(self.wb, daughter=False,name='ук без дочерними')

        print(bcolors.OKGREEN + '\rДва укашных файла созданы' + bcolors.ENDC)

    def without_uk(self):
        print('\rДелаем файл без ук', end='')

        s = self.wb['Sheet']

        wb_daughters, wb_without_daughter = Workbook(), Workbook()
        s_d, s_without = wb_daughters.active, wb_without_daughter.active

        r_d, r_without = 1, 1
        i = 2

        while True:
            checker = s.cell(row=i, column=13).value == None and s.cell(row=i, column=19).value == None
            if checker:
                break

            if not (s.cell(row=i, column=1).value is None and s.cell(row=i, column=13).value is not None):
                i += 1
                continue

            for col_ in range(1, 30):
                s_d.cell(row=r_d, column=col_).value = s.cell(row=i, column=col_).value
                s_without.cell(row=r_without, column=col_).value = s.cell(row=i, column=col_).value

            r_d += 1
            r_without += 1
            i += 1

            while s.cell(row=i, column=14).value is None:
                for col_ in range(19, 50):
                    s_d.cell(row=r_d, column=col_).value = s.cell(row=i, column=col_).value
                r_d += 1
                i += 1
                if s.cell(row=i, column=19).value == None:
                    break

        wb_without_daughter.save('Выгрузка/Адреса без ук, без дочерних.xlsx')
        wb_daughters.save('Выгрузка/Адреса без ук, с дочерними.xlsx')

        print(bcolors.OKGREEN + '\rДва файла без ук созданы' + bcolors.ENDC)

