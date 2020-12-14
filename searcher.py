import copy
import glob
import json

import psycopg2
from openpyxl import load_workbook, Workbook

import info
from colors import bcolors


class Search():
    def __init__(self, file_name):
        self.file_name = file_name
        self.conn = psycopg2.connect(dbname='reimport2', user='cuba',
                                     password='cuba', host='localhost')
        self.cursor = self.conn.cursor()
        # self.wb = load_workbook('Выгрузка/1.xlsx') #Для теста

    def put_daughter(self):
        print('\rВставляем дочерние кадастры', end='')

        wb = load_workbook('Выгрузка/tmp/tmp.xlsx')
        s = wb['Sheet']

        wb_new = Workbook()
        s_new = wb_new.active
        row_new = 2

        apartments = {}
        owners = {}
        for apartment, list_daughter in info.a[0]:
            if len(list_daughter) > 0:
                apartments[apartment] = list_daughter

        for apartment, list_owners in info.a[2]:
            if len(list_owners) > 0:
                owners[apartment] = list_owners

        self.apartments = apartments
        self.owners = owners

        for i in range(1, 60):
            s_new.cell(row=1, column=i).value = s.cell(row=1, column=i).value
        for i in range(2, 500000):
            kn = s.cell(row=i, column=13).value
            if kn == None:
                break
            if kn not in apartments:
                continue

            for j in range(1, 30):
                s_new.cell(row=row_new, column=j).value = s.cell(row=i, column=j).value
            s_new.cell(row=row_new, column=18).value = len(apartments[kn])
            row_new += 1
            for daugther in apartments[kn]:
                s_new.cell(row=row_new, column=19).value = daugther
                row_new += 1
        self.wb = wb_new

        print(bcolors.OKGREEN + '\r2) Сформирован файл со всеми дочерними адресами' + bcolors.ENDC)

    def put_info(self, cash_size):
        print('\rВставляем информацию о дочерних кадастрах: {0}%'.format(0), end='')
        s = self.wb['Sheet']
        cnt = 0
        for i in range(2, 500000):
            if s.cell(row=i, column=18).value == None and s.cell(row=i, column=19).value == None:
                cnt = i
                break
        self.cnt = cnt
        list_kn = []
        for i in range(2, cnt):
            print('\rВставляем информацию о дочерних кадастрах: {0}% ({1}/{2})'.format(i * 100 // cnt, i, cnt), end='')
            type = s.cell(row=i, column=18).value != None
            if not type:
                kn = s.cell(row=i, column=19).value
                list_kn.append((i, kn))

            if len(list_kn) >= cash_size or i == cnt - 1:
                query = "select * from reimport_rtneo_refactor where cadastral_number in ("
                id_by_kn = {}
                for id, kn in list_kn:
                    query += "'{0}',".format(kn)
                    id_by_kn[kn] = id
                list_kn = []
                query = query[:-1] + ")"
                self.cursor.execute(query)
                responses = self.cursor.fetchall()
                for response in responses:
                    row = id_by_kn[response[6]]
                    kn = response[6]
                    for col, res in [(20, 3), (21, 4), (22, 5), (23, 7), (24, 8), (25, 9), (26, 10)]:
                        s.cell(row=row, column=col).value = response[res]

                    if kn in self.owners:
                        for id, owner in enumerate(self.owners[kn]):
                            s.cell(row=row, column=27 + id).value = owner
        print('\rСохраняем', end='')
        self.wb.save('Выгрузка/1.xlsx')
        print(bcolors.OKGREEN + '\r3) Информация вставлена в файл с дочерними кадастрами' + bcolors.ENDC)

    def reformat_uk_json(self):
        print('\rОбработка uk_JSON', end='')
        files = glob.glob('Выгрузка/uk_json/*.txt')
        type_of, super_dict = {}, {}

        for f in files:
            dic = open(f, 'r').read()
            dic = json.loads(dic)
            type_of[f.split('.')[0]] = dic

        for key, val in type_of.items():
            for i in val['House']:
                house = i['Address'].split(', ')[-1].split('. ')[1].upper()
                street = i['Address'].split(', ')[-2].split('. ')[1].upper()
                super_dict['{0}||{1}'.format(street, house)] = {'type uk': key, 'LicenseNumber': val['LicenseNumber'],
                                                                'LicenseRegDate': val['LicenseRegDate'], 'house': i}

        f = open('Выгрузка/uk_json/super_dict.json', 'w', encoding='utf-8')
        json.dump(super_dict, f, ensure_ascii=False)
        print(bcolors.OKGREEN + '\r4) Словарь с обработанными адресами uk создан' + bcolors.ENDC)

    def put_uk_info(self):
        print('\rВставляем информацию из ук в файл с дочерними', end='')
        s = self.wb['Sheet']
        inf = json.loads(open('Выгрузка/uk_json/super_dict.json', 'r').read())
        not_found_dict = copy.deepcopy(inf)

        for i in range(2, self.cnt):
            street = s.cell(row=i, column=8).value
            house = s.cell(row=i, column=9).value
            key = '{0}||{1}'.format(street, house)

            if key in inf:

                s.cell(row=i, column=1).value = inf[key].get('type uk')
                s.cell(row=i, column=2).value = inf[key].get('LicenseNumber')
                s.cell(row=i, column=3).value = inf[key].get('LicenseRegDate')
                s.cell(row=i, column=4).value = inf[key]['house'].get('FiasHouseGuid')
                s.cell(row=i, column=5).value = inf[key]['house'].get('ContractGuid')

                if key in not_found_dict:
                    del not_found_dict[key]

        self.wb.save('Выгрузка/2.xlsx')
        f = open('Выгрузка/непопавшие.json', 'w', encoding='utf-8')
        json.dump(not_found_dict, f, ensure_ascii=False, sort_keys=True, indent=4, )

        print(bcolors.OKGREEN + '\r5) Информация из ук была проставлена. Не синхронизировалось {0} адресов'.format(
            len(not_found_dict)) + bcolors.ENDC)
